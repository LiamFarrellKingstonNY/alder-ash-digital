from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

FOREST_GREEN = "254D3B"
CREAM = "F7F2E8"
AMBER = "C4892A"
WHITE = "FFFFFF"
LIGHT_GREEN = "D6E8DF"
BLUE_INPUT = "0000FF"
BLACK = "000000"
GREEN_LINK = "008000"

FONT_NAME = "Arial"

CURRENCY_FMT = '$#,##0;($#,##0);"-"'
CURRENCY_DEC_FMT = '$#,##0.00;($#,##0.00);"-"'
PCT_FMT = '0.0%;-0.0%;"-"'
NUMBER_FMT = '#,##0;(#,##0);"-"'

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return PatternFill("solid", start_color=FOREST_GREEN, end_color=FOREST_GREEN)

def cream_fill():
    return PatternFill("solid", start_color=CREAM, end_color=CREAM)

def amber_fill():
    return PatternFill("solid", start_color=AMBER, end_color=AMBER)

def white_fill():
    return PatternFill("solid", start_color=WHITE, end_color=WHITE)

def light_green_fill():
    return PatternFill("solid", start_color=LIGHT_GREEN, end_color=LIGHT_GREEN)

def hdr_font(bold=True, size=10):
    return Font(name=FONT_NAME, bold=bold, color=WHITE, size=size)

def body_font(bold=False, color=BLACK, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def amber_font(bold=True):
    return Font(name=FONT_NAME, bold=bold, color=WHITE, size=10)

def style_cell(cell, value=None, font=None, fill=None, fmt=None,
               align_h='center', align_v='center', border=True, wrap=False):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v,
                                wrap_text=wrap)
    if border:
        cell.border = thin_border()

def build_monthly_sheet(wb):
    ws = wb.create_sheet("Monthly Projections")
    ws.sheet_view.showGridLines = False

    months = ["Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26",
              "Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26"]

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    style_cell(c, "ALDER & ASH DIGITAL — Year 1 Monthly Projections (2026)",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=13),
               fill=header_fill(), align_h="center", border=False)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"].value = ""
    ws["A2"].fill = white_fill()

    # ── Column headers ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws["A3"].value = "Metric"
    style_cell(ws["A3"], font=hdr_font(), fill=header_fill(), align_h="left")

    for i, m in enumerate(months):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 12
        style_cell(ws[f"{col}3"], m, font=hdr_font(), fill=header_fill())

    col_N = "N"
    ws.column_dimensions[col_N].width = 14
    style_cell(ws[f"{col_N}3"], "FY 2026", font=hdr_font(), fill=amber_fill())
    ws.row_dimensions[3].height = 20

    # ── Row definitions ─────────────────────────────────────────────────────
    # projects ramp: 1,1,1,2,2,3,3,3,3,3,3,3
    project_counts = [1,1,1,2,2,3,3,3,3,3,3,3]
    avg_val = 3500
    retainer_rate = 250
    sw_cost = 200
    mkt_cost = 300
    ins_cost = 100
    misc_cost = 150

    # Row map: row_number -> (label, section)
    rows = [
        (4,  "── REVENUE ──",          "section"),
        (5,  "Website Projects Closed", "input"),
        (6,  "Avg Project Value ($)",   "input"),
        (7,  "Project Revenue",         "formula"),
        (8,  "New Retainer Clients",    "formula"),
        (9,  "Cumulative Retainers",    "formula"),
        (10, "Retainer Rate ($/mo)",    "input"),
        (11, "Retainer Revenue",        "formula"),
        (12, "Add-On Revenue (30%)",    "formula"),
        (13, "TOTAL MONTHLY REVENUE",   "total"),
        (14, "── EXPENSES ──",          "section"),
        (15, "Software & Tools",        "input"),
        (16, "Marketing",               "input"),
        (17, "Insurance",               "input"),
        (18, "Miscellaneous",           "input"),
        (19, "TOTAL EXPENSES",          "total"),
        (20, "── PROFIT ──",            "section"),
        (21, "NET PROFIT",              "key"),
        (22, "CUMULATIVE REVENUE",      "key"),
        (23, "PROFIT MARGIN",           "key_pct"),
    ]

    section_labels = {"── REVENUE ──","── EXPENSES ──","── PROFIT ──"}

    def row_fill(r, row_type, col_is_total=False):
        if row_type == "section":
            return PatternFill("solid", start_color=FOREST_GREEN, end_color=FOREST_GREEN)
        if row_type in ("total","key","key_pct"):
            return amber_fill()
        if col_is_total:
            return amber_fill()
        if r % 2 == 0:
            return cream_fill()
        return white_fill()

    def row_font(row_type, col_is_total=False, is_input=False):
        if row_type == "section":
            return Font(name=FONT_NAME, bold=True, color=WHITE, size=9, italic=True)
        if row_type in ("total","key","key_pct") or col_is_total:
            return Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
        if is_input:
            return Font(name=FONT_NAME, color=BLUE_INPUT, size=10)
        return body_font()

    for row_num, label, row_type in rows:
        ws.row_dimensions[row_num].height = 18

        # Label cell
        c = ws[f"A{row_num}"]
        is_section = row_type == "section"
        style_cell(c, label,
                   font=row_font(row_type),
                   fill=row_fill(row_num, row_type),
                   align_h="left" if not is_section else "center",
                   border=not is_section)
        if is_section:
            ws.merge_cells(f"A{row_num}:N{row_num}")
            continue

        # Data columns B..M
        for i, m in enumerate(months):
            col_idx = i + 2
            col = get_column_letter(col_idx)
            prev_col = get_column_letter(col_idx - 1)
            cell = ws[f"{col}{row_num}"]
            is_total_col = False
            fmt = CURRENCY_FMT
            val = None

            if row_type == "input":
                if label == "Website Projects Closed":
                    val = project_counts[i]
                    fmt = NUMBER_FMT
                elif label == "Avg Project Value ($)":
                    val = avg_val
                elif label == "Retainer Rate ($/mo)":
                    val = retainer_rate
                elif label == "Software & Tools":
                    val = sw_cost
                elif label == "Marketing":
                    val = mkt_cost
                elif label == "Insurance":
                    val = ins_cost
                elif label == "Miscellaneous":
                    val = misc_cost

            elif row_type == "formula":
                if label == "Project Revenue":
                    val = f"={col}5*{col}6"
                elif label == "New Retainer Clients":
                    val = f"={col}5"
                    fmt = NUMBER_FMT
                elif label == "Cumulative Retainers":
                    if col_idx == 2:
                        val = f"={col}8"
                    else:
                        val = f"={prev_col}9+{col}8"
                    fmt = NUMBER_FMT
                elif label == "Retainer Revenue":
                    val = f"={col}9*{col}10"
                elif label == "Add-On Revenue (30%)":
                    val = f"={col}7*0.3"

            elif row_type == "total":
                if label == "TOTAL MONTHLY REVENUE":
                    val = f"={col}7+{col}11+{col}12"
                elif label == "TOTAL EXPENSES":
                    val = f"=SUM({col}15:{col}18)"

            elif row_type == "key":
                if label == "NET PROFIT":
                    val = f"={col}13-{col}19"
                elif label == "CUMULATIVE REVENUE":
                    if col_idx == 2:
                        val = f"={col}13"
                    else:
                        val = f"={prev_col}22+{col}13"

            elif row_type == "key_pct":
                if label == "PROFIT MARGIN":
                    val = f"=IF({col}13=0,0,{col}21/{col}13)"
                    fmt = PCT_FMT

            style_cell(cell,
                       value=val,
                       font=row_font(row_type, is_total_col, row_type=="input"),
                       fill=row_fill(row_num, row_type, is_total_col),
                       fmt=fmt)

        # FY Total column N
        n_cell = ws[f"N{row_num}"]
        n_fmt = CURRENCY_FMT
        n_val = None

        if label == "Website Projects Closed":
            n_val = "=SUM(B5:M5)"; n_fmt = NUMBER_FMT
        elif label == "Avg Project Value ($)":
            n_val = "=AVERAGE(B6:M6)"
        elif label == "Project Revenue":
            n_val = "=SUM(B7:M7)"
        elif label == "New Retainer Clients":
            n_val = "=SUM(B8:M8)"; n_fmt = NUMBER_FMT
        elif label == "Cumulative Retainers":
            n_val = "=M9"; n_fmt = NUMBER_FMT
        elif label == "Retainer Rate ($/mo)":
            n_val = "=AVERAGE(B10:M10)"
        elif label == "Retainer Revenue":
            n_val = "=SUM(B11:M11)"
        elif label == "Add-On Revenue (30%)":
            n_val = "=SUM(B12:M12)"
        elif label == "TOTAL MONTHLY REVENUE":
            n_val = "=SUM(B13:M13)"
        elif label == "Software & Tools":
            n_val = "=SUM(B15:M15)"
        elif label == "Marketing":
            n_val = "=SUM(B16:M16)"
        elif label == "Insurance":
            n_val = "=SUM(B17:M17)"
        elif label == "Miscellaneous":
            n_val = "=SUM(B18:M18)"
        elif label == "TOTAL EXPENSES":
            n_val = "=SUM(B19:M19)"
        elif label == "NET PROFIT":
            n_val = "=SUM(B21:M21)"
        elif label == "CUMULATIVE REVENUE":
            n_val = "=M22"
        elif label == "PROFIT MARGIN":
            n_val = "=IF(N13=0,0,N21/N13)"; n_fmt = PCT_FMT

        style_cell(n_cell, value=n_val,
                   font=Font(name=FONT_NAME, bold=True, color=WHITE, size=10),
                   fill=amber_fill(), fmt=n_fmt)

    # ── Notes row ────────────────────────────────────────────────────────────
    ws.row_dimensions[24].height = 14
    ws.merge_cells("A24:N24")
    nc = ws["A24"]
    nc.value = "Source: Internal projections, Alder & Ash Digital business plan. Blue cells = inputs (editable). All calculations via Excel formulas."
    nc.font = Font(name=FONT_NAME, size=8, italic=True, color="888888")
    nc.alignment = Alignment(horizontal="left", vertical="center")
    nc.fill = white_fill()

    ws.freeze_panes = "B4"
    return ws


def build_annual_sheet(wb):
    ws = wb.create_sheet("Annual Summary")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Title
    ws.merge_cells("A1:D1")
    style_cell(ws["A1"], "ALDER & ASH DIGITAL — Annual Summary",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=13),
               fill=header_fill(), align_h="center", border=False)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws["A2"].fill = white_fill()

    # Column headers
    for col, label, fill in [("A","Metric",header_fill()),
                               ("B","Year 1 (2026)",header_fill()),
                               ("C","Year 2 (2027)",header_fill()),
                               ("D","YoY Change",amber_fill())]:
        style_cell(ws[f"{col}3"], label, font=hdr_font(), fill=fill)
    ws.row_dimensions[3].height = 20

    mp = "'Monthly Projections'"

    # Y2 assumptions (hardcoded inputs, blue)
    y2_revenue     = 185000   # estimate
    y2_expenses    = 18000    # estimate (annual)
    y2_retainers   = 28       # end-of-year retainers
    y2_projects    = 42       # sites completed
    y2_mrr         = 7000     # MRR from retainers

    data_rows = [
        # (row, label, y1_formula, y2_value, fmt, row_type)
        (4,  "── REVENUE METRICS ──",         None, None, None, "section"),
        (5,  "Total Revenue ($)",              f"={mp}!N13", y2_revenue, CURRENCY_FMT, "formula"),
        (6,  "Total Expenses ($)",             f"={mp}!N19", y2_expenses, CURRENCY_FMT, "formula"),
        (7,  "Net Profit ($)",                 f"={mp}!N21", None,       CURRENCY_FMT, "key"),
        (8,  "Profit Margin (%)",              f"={mp}!N23", None,       PCT_FMT,      "key_pct"),
        (9,  "── VOLUME METRICS ──",           None, None, None, "section"),
        (10, "Projects Completed",             f"={mp}!N5",  y2_projects, NUMBER_FMT, "formula"),
        (11, "Avg Projects/Month",             f"={mp}!N5/12", y2_projects/12, NUMBER_FMT, "formula"),
        (12, "End-of-Year Retainers",          f"={mp}!N9",  y2_retainers, NUMBER_FMT, "formula"),
        (13, "── MRR & GROWTH ──",             None, None, None, "section"),
        (14, "MRR from Retainers (Dec)",       f"={mp}!M11", y2_mrr,    CURRENCY_FMT, "formula"),
        (15, "Avg Monthly Revenue",            f"={mp}!N13/12", None,  CURRENCY_FMT, "key"),
        (16, "Avg Monthly Expenses",           f"={mp}!N19/12", None,  CURRENCY_FMT, "key"),
        (17, "── YEAR 2 ASSUMPTIONS ──",       None, None, None, "section"),
        (18, "Target Sites/Month (avg)",       None, 3.5,  NUMBER_FMT, "input"),
        (19, "Target Retainers (EoY)",         None, 28,   NUMBER_FMT, "input"),
        (20, "Avg Project Value ($)",          None, 4000, CURRENCY_FMT, "input"),
        (21, "Retainer Rate ($/mo)",           None, 275,  CURRENCY_FMT, "input"),
    ]

    for row_num, label, y1_val, y2_val, fmt, row_type in data_rows:
        ws.row_dimensions[row_num].height = 18
        is_section = row_type == "section"

        fill_a = header_fill() if is_section else (
            amber_fill() if row_type in ("key","key_pct") else
            (cream_fill() if row_num % 2 == 0 else white_fill())
        )
        font_a = (Font(name=FONT_NAME, bold=True, color=WHITE, size=9, italic=True) if is_section
                  else (Font(name=FONT_NAME, bold=True, color=WHITE, size=10) if row_type in ("key","key_pct")
                        else body_font()))

        style_cell(ws[f"A{row_num}"], label, font=font_a, fill=fill_a,
                   align_h="left" if not is_section else "center", border=not is_section)
        if is_section:
            ws.merge_cells(f"A{row_num}:D{row_num}")
            continue

        # Y1
        b = ws[f"B{row_num}"]
        b_fill = amber_fill() if row_type in ("key","key_pct") else fill_a
        b_font = (Font(name=FONT_NAME, bold=True, color=WHITE, size=10)
                  if row_type in ("key","key_pct") else
                  Font(name=FONT_NAME, color=GREEN_LINK, size=10))
        if row_type in ("key","key_pct") and y1_val:
            # derive from Y1
            if row_num == 7:
                y1_val = f"=B5-B6"
            elif row_num == 8:
                y1_val = f"=IF(B5=0,0,B7/B5)"
            elif row_num == 15:
                y1_val = f"=B5/12"
            elif row_num == 16:
                y1_val = f"=B6/12"
        style_cell(b, value=y1_val, font=b_font, fill=b_fill, fmt=fmt)

        # Y2
        c = ws[f"C{row_num}"]
        c_fill = amber_fill() if row_type in ("key","key_pct") else fill_a
        c_font_color = BLUE_INPUT if row_type == "input" else (
            WHITE if row_type in ("key","key_pct") else BLACK)
        c_font = Font(name=FONT_NAME, bold=(row_type in ("key","key_pct")),
                      color=c_font_color, size=10)

        if row_type == "key" and row_num == 7:
            y2_val = f"=C5-C6"
        elif row_type == "key_pct" and row_num == 8:
            y2_val = f"=IF(C5=0,0,C7/C5)"
        elif row_type == "key" and row_num == 15:
            y2_val = f"=C5/12"
        elif row_type == "key" and row_num == 16:
            y2_val = f"=C6/12"

        style_cell(c, value=y2_val, font=c_font, fill=c_fill, fmt=fmt)

        # YoY
        d = ws[f"D{row_num}"]
        d_fill = amber_fill() if row_type in ("key","key_pct") else fill_a
        d_font = Font(name=FONT_NAME, bold=True, color=WHITE if row_type in ("key","key_pct") else BLACK, size=10)
        if fmt == PCT_FMT:
            d_val = f"=IF(B{row_num}=0,0,C{row_num}-B{row_num})"
            d_fmt = "+0.0%;-0.0%;-"
        else:
            d_val = f"=IF(B{row_num}=0,0,(C{row_num}-B{row_num})/B{row_num})"
            d_fmt = "+0.0%;-0.0%;-"
        style_cell(d, value=d_val, font=d_font, fill=d_fill, fmt=d_fmt)

    ws.row_dimensions[22].height = 14
    ws.merge_cells("A22:D22")
    nc = ws["A22"]
    nc.value = ("Year 2 figures are management estimates. Blue cells = editable inputs. "
                "Green cells = linked from Monthly Projections sheet.")
    nc.font = Font(name=FONT_NAME, size=8, italic=True, color="888888")
    nc.alignment = Alignment(horizontal="left", vertical="center")
    nc.fill = white_fill()

    ws.freeze_panes = "B4"
    return ws


def build_pricing_sheet(wb):
    ws = wb.create_sheet("Pricing Calculator")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20

    ws.merge_cells("A1:E1")
    style_cell(ws["A1"], "ALDER & ASH DIGITAL — Pricing Calculator",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=13),
               fill=header_fill(), align_h="center", border=False)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2"); ws["A2"].fill = white_fill()

    # ── INPUT SECTION ──
    ws.merge_cells("A3:B3")
    style_cell(ws["A3"], "INPUTS — Edit Blue Cells",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=10),
               fill=header_fill(), align_h="center")
    ws["B3"].fill = header_fill()

    inputs = [
        (4,  "Project Price ($)",          5000,  CURRENCY_FMT),
        (5,  "Retainer Price ($/mo)",       250,   CURRENCY_FMT),
        (6,  "Projects per Month",           2,    NUMBER_FMT),
        (7,  "Client Retention Rate (%)",   0.75,  PCT_FMT),
        (8,  "Add-On % of Project Rev",     0.30,  PCT_FMT),
        (9,  "Monthly Fixed Expenses ($)",   750,  CURRENCY_FMT),
    ]

    for row, label, val, fmt in inputs:
        ws.row_dimensions[row].height = 18
        a = ws[f"A{row}"]
        b = ws[f"B{row}"]
        fill = cream_fill() if row % 2 == 0 else white_fill()
        style_cell(a, label, font=body_font(), fill=fill, align_h="left")
        style_cell(b, val,
                   font=Font(name=FONT_NAME, color=BLUE_INPUT, bold=True, size=10),
                   fill=PatternFill("solid", start_color="EAF4FF", end_color="EAF4FF"),
                   fmt=fmt)

    # ── OUTPUT SECTION ──
    ws.merge_cells("A11:B11")
    style_cell(ws["A11"], "PROJECTED OUTPUTS",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=10),
               fill=amber_fill(), align_h="center")
    ws["B11"].fill = amber_fill()

    # Monthly breakdown header
    ws.merge_cells("A12:B12")
    style_cell(ws["A12"], "Monthly Breakdown",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=9),
               fill=header_fill(), align_h="center")
    ws["B12"].fill = header_fill()

    monthly_outputs = [
        (13, "Monthly Project Revenue",   "=B6*B4",                     CURRENCY_FMT),
        (14, "Monthly Retainer Revenue",  "=B6*B7*B5",                  CURRENCY_FMT),
        (15, "Monthly Add-On Revenue",    "=B13*B8",                    CURRENCY_FMT),
        (16, "Total Monthly Revenue",     "=B13+B14+B15",               CURRENCY_FMT),
        (17, "Total Monthly Expenses",    "=B9",                        CURRENCY_FMT),
        (18, "Monthly Net Profit",        "=B16-B17",                   CURRENCY_FMT),
        (19, "Monthly Profit Margin",     "=IF(B16=0,0,B18/B16)",       PCT_FMT),
    ]

    for row, label, formula, fmt in monthly_outputs:
        ws.row_dimensions[row].height = 18
        fill = amber_fill() if row in (16, 18, 19) else (cream_fill() if row % 2 == 0 else white_fill())
        font_a = Font(name=FONT_NAME, bold=(row in (16,18,19)),
                      color=WHITE if row in (16,18,19) else BLACK, size=10)
        font_b = Font(name=FONT_NAME, bold=(row in (16,18,19)),
                      color=WHITE if row in (16,18,19) else BLACK, size=10)
        style_cell(ws[f"A{row}"], label, font=font_a, fill=fill, align_h="left")
        style_cell(ws[f"B{row}"], formula, font=font_b, fill=fill, fmt=fmt)

    # Annual outputs
    ws.merge_cells("A21:B21")
    style_cell(ws["A21"], "Annual Projection",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=9),
               fill=header_fill(), align_h="center")
    ws["B21"].fill = header_fill()

    annual_outputs = [
        (22, "Annual Project Revenue",    "=B13*12",   CURRENCY_FMT),
        (23, "Annual Retainer Revenue",   "=B14*12",   CURRENCY_FMT),
        (24, "Annual Add-On Revenue",     "=B15*12",   CURRENCY_FMT),
        (25, "Total Annual Revenue",      "=B16*12",   CURRENCY_FMT),
        (26, "Total Annual Expenses",     "=B17*12",   CURRENCY_FMT),
        (27, "Annual Net Profit",         "=B25-B26",  CURRENCY_FMT),
        (28, "Annual Profit Margin",      "=IF(B25=0,0,B27/B25)", PCT_FMT),
    ]

    for row, label, formula, fmt in annual_outputs:
        ws.row_dimensions[row].height = 18
        fill = amber_fill() if row in (25, 27, 28) else (cream_fill() if row % 2 == 0 else white_fill())
        font_color = WHITE if row in (25,27,28) else BLACK
        style_cell(ws[f"A{row}"], label,
                   font=Font(name=FONT_NAME, bold=(row in (25,27,28)), color=font_color, size=10),
                   fill=fill, align_h="left")
        style_cell(ws[f"B{row}"], formula,
                   font=Font(name=FONT_NAME, bold=(row in (25,27,28)), color=font_color, size=10),
                   fill=fill, fmt=fmt)

    # Break-even analysis
    ws.merge_cells("A30:B30")
    style_cell(ws["A30"], "Break-Even Analysis",
               font=Font(name=FONT_NAME, bold=True, color=WHITE, size=9),
               fill=header_fill(), align_h="center")
    ws["B30"].fill = header_fill()

    be_rows = [
        (31, "Monthly Fixed Costs",           "=B9",                              CURRENCY_FMT),
        (32, "Revenue per Project",            "=B4*(1+B8)",                      CURRENCY_FMT),
        (33, "Break-Even Projects/Month",      "=IF(B32=0,0,B31/B32)",           "#,##0.0;(#,##0.0);-"),
        (34, "Break-Even Revenue/Month",       "=B31",                            CURRENCY_FMT),
        (35, "Safety Margin",                  "=IF(B16=0,0,(B16-B31)/B16)",     PCT_FMT),
    ]

    for row, label, formula, fmt in be_rows:
        ws.row_dimensions[row].height = 18
        fill = light_green_fill() if row in (33,35) else (cream_fill() if row % 2 == 0 else white_fill())
        style_cell(ws[f"A{row}"], label,
                   font=Font(name=FONT_NAME, bold=(row in (33,35)), color=BLACK, size=10),
                   fill=fill, align_h="left")
        style_cell(ws[f"B{row}"], formula,
                   font=Font(name=FONT_NAME, bold=(row in (33,35)), color=BLACK, size=10),
                   fill=fill, fmt=fmt)

    ws.row_dimensions[37].height = 14
    ws.merge_cells("A37:E37")
    nc = ws["A37"]
    nc.value = "Blue cells = inputs. Green cells = break-even outputs. All outputs recalculate automatically when inputs change."
    nc.font = Font(name=FONT_NAME, size=8, italic=True, color="888888")
    nc.alignment = Alignment(horizontal="left", vertical="center")
    nc.fill = white_fill()

    ws.freeze_panes = "A11"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_monthly_sheet(wb)
    build_annual_sheet(wb)
    build_pricing_sheet(wb)

    out = "/Users/ashai/Desktop/alder-ash-digital/Alder-Ash-Financials.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
